import os
import asyncio
import datetime
import requests
import uuid
import re
from collections import defaultdict

from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import (
    Application, CommandHandler, MessageHandler,
    CallbackQueryHandler, filters, ContextTypes
)
from openpyxl import Workbook, load_workbook
from bs4 import BeautifulSoup

# Токен вашего бота
TOKEN = os.getenv("TELEGRAM_BOT_TOKEN1")
XLSX_FILE = "uploads.xlsx"

# Буферы и константы
pending_bytes = defaultdict(list)  # Для байтов фото по (chat_id, media_group_id или file_id)
recent_uploads = {}  # Кэш последних загруженных файлов: file_id -> datetime
choice_keys = {}  # Для хранения коротких ID -> реального key
DUPLICATE_INTERVAL = datetime.timedelta(minutes=10)

def ensure_workbook():
    if not os.path.exists(XLSX_FILE):
        wb = Workbook()
        ws = wb.active
        ws.append(["Timestamp", "UserID", "Username", "FileID", "URL", "Hosting"])
        wb.save(XLSX_FILE)

def append_record(user_id: int, username: str, file_id: str, url: str, hosting: str):
    wb = load_workbook(XLSX_FILE)
    ws = wb.active
    timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws.append([timestamp, user_id, username or "", file_id, url, hosting])
    wb.save(XLSX_FILE)

def upload_to_anoimage(image_bytes: bytes, filename: str) -> str:
    files = {"file": (filename, image_bytes, "image/jpeg")}
    headers = {
        "Origin": "https://anoimage.com",
        "Referer": "https://anoimage.com/",
        "X-Requested-With": "XMLHttpRequest"
    }
    resp = requests.post("https://anoimage.com/upload-image.php", files=files, headers=headers, timeout=30)
    resp.raise_for_status()
    data = resp.json()
    image_id = next((v for k, v in data.items() if k.isdigit()), None)
    if not image_id:
        raise ValueError(f"Неверный формат ответа: {data}")
    return f"https://anoimage.com/{image_id}"

def upload_to_ninjabox(image_bytes: bytes, filename: str) -> str:
    files = {"files": (filename, image_bytes, "image/jpeg")}
    data = {"password": ""}
    headers = {
        "Origin": "https://ninjabox.org",
        "Referer": "https://ninjabox.org/",
        "User-Agent": "Mozilla/5.0"
    }
    resp = requests.post("https://ninjabox.org/put", files=files, data=data,
                         headers=headers, timeout=40, allow_redirects=True)
    resp.raise_for_status()
    response_text = resp.text
    if "main-form" in response_text or "files to upload" in response_text:
        raise ValueError("Загрузка не удалась: вернулась главная страница.")
    soup = BeautifulSoup(response_text, "html.parser")
    possible_selectors = [
        soup.find("input", {"class": "share-input"}),
        soup.find("input", {"id": "share-link"}),
        soup.find("input", {"class": "link-input"}),
        soup.find("a", {"class": "share-link"}),
        soup.find("a", {"class": "direct-link"}),
        soup.find("input", {"type": "text", "readonly": "readonly"}),
        soup.find("div", {"class": "share-block"}),
        soup.find("input", {"name": "link"})
    ]
    for element in possible_selectors:
        if element:
            url = element.get("value") or element.get("href") or element.text.strip()
            if url and "nbox.me" in url:
                return url
    url_match = re.search(r"https://nbox\.me/[a-f0-9\-]+", response_text)
    if url_match:
        return url_match.group(0)
    raise ValueError("Не удалось извлечь ссылку. Структура страницы не соответствует ожидаемой.")

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    sticker = "CAACAgIAAxkBAAEPJWRonJ6NS7DK4cSC8GBQ768xBoZG1wACDwEAAlKJkSNldRdchg_VhjYE"
    await context.bot.send_sticker(update.effective_chat.id, sticker)
    text = (
        "<b>👋 Конничива, бро и лёгкой руки!</b>\n\n"
        "Отправь фото — я загружу на выбранный хостинг быстро и анонимно 🥷🏻."
    )
    await update.message.reply_html(
        text,
        reply_markup=InlineKeyboardMarkup([
            [InlineKeyboardButton("ℹ️ Почему мы?🥷🏻", callback_data="about")],
            [InlineKeyboardButton("📤 Загрузить фото 🥷🏻", callback_data="upload")]
        ])
    )

async def help_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_html(
        "<b>ℹ️ Команды бота:</b>\n"
        "/start — запустить бота\n"
        "/help — показать справку"
    )

async def process_pending(key, hosting, context, msg, user, now):
    await asyncio.sleep(1)
    photos = pending_bytes.pop(key, [])
    urls = []
    for file_id, img_bytes in photos:
        try:
            if hosting == "anoimage":
                url = upload_to_anoimage(img_bytes, f"{file_id}.jpg")
            else:
                url = upload_to_ninjabox(img_bytes, f"{file_id}.jpg")
            append_record(user.id, user.username, file_id, url, hosting)
            recent_uploads[file_id] = now
            urls.append(url)
        except Exception as e:
            urls.append(f"Ошибка: {e}")
    if len(urls) > 1:
        lines = [f"{i+1}: {u}" for i, u in enumerate(urls)]
        text = f"<b>✅ Мультизагрузка на {hosting.capitalize()} завершена:</b>\n" + "\n".join(lines)
    else:
        text = f"✅ <b>Лови адрес на {hosting.capitalize()}:</b>\n{urls[0]}"
    reply_markup = InlineKeyboardMarkup([[InlineKeyboardButton("🔙 Назад к выбору", callback_data="upload")]])
    await msg.reply_html(text, disable_web_page_preview=True, reply_markup=reply_markup)

async def on_photo(update: Update, context: ContextTypes.DEFAULT_TYPE):
    msg = update.message
    user = update.effective_user
    photo = msg.photo[-1]
    file_id = photo.file_id
    now = datetime.datetime.now()
    mgid = msg.media_group_id
    key = (msg.chat.id, mgid) if mgid else (msg.chat.id, file_id)
    last = recent_uploads.get(file_id)
    if last and now - last < DUPLICATE_INTERVAL:
        await msg.reply_html("❗️ Бро, ты грузанул где-то дубль, будь внимателен.")
        return
    img_bytes = await (await context.bot.get_file(file_id)).download_as_bytearray()
    pending_bytes[key].append((file_id, img_bytes))
    hosting = context.user_data.get("selected_hosting")
    if hosting:
        if len(pending_bytes[key]) == 1 and mgid:
            context.application.create_task(process_pending(key, hosting, context, msg, user, now))
        elif not mgid:
            await process_pending(key, hosting, context, msg, user, now)
    else:
        short_id = str(uuid.uuid4())[:8]
        choice_keys[short_id] = key
        buttons = [
            [InlineKeyboardButton("Anoimage.com (120 дней)", callback_data=f"host_ano_{short_id}")],
            [InlineKeyboardButton("Ninjabox.org (180 дней)", callback_data=f"host_ninja_{short_id}")]
        ]
        await msg.reply_html("<b>Выберите хостинг для загрузки:</b>", reply_markup=InlineKeyboardMarkup(buttons))

async def button_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    data = q.data
    if data.startswith("host_"):
        parts = data.split("_")
        hosting = "anoimage" if parts[1] == "ano" else "ninjabox"
        short_id = parts[2]
        key = choice_keys.get(short_id)
        choice_keys.pop(short_id, None)
        context.user_data["selected_hosting"] = hosting
        if key:
            await on_photo(update, context)
        else:
            days = "120" if hosting == "anoimage" else "180"
            await q.edit_message_text(
                f"Выбран {hosting.capitalize()} (хранение: {days} дней).\n"
                "📸 Теперь пришли мне фотку или фотки (максимум 10 штук за один раз) для загрузки.",
                parse_mode="HTML",
                reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🔄 Сменить хостинг", callback_data="upload")]])
            )
    elif data == "upload":
        short_id = str(uuid.uuid4())[:8]
        buttons = [
            [InlineKeyboardButton("Anoimage.com (120 дней)", callback_data=f"host_ano_{short_id}")],
            [InlineKeyboardButton("Ninjabox.org (180 дней)", callback_data=f"host_ninja_{short_id}")]
        ]
        await q.edit_message_text(
            "<b>Выбери хостинг, который тебе нравится, бро:</b>",
            parse_mode="HTML",
            reply_markup=InlineKeyboardMarkup(buttons)
        )
    elif data == "about":
        about = (
            "<b>🥷🏻Почему мы?🥷🏻</b>\n\n"
            "1️⃣ Не сохраняем файлы на наших серверах.\n"
            "2️⃣ Не требуем регистрации.\n"
            "3️⃣ Конфиденциальность — без логов.\n"
            "4️⃣ Автоудаление: 120 дней (Anoimage) или 180 дней (Ninjabox).\n"
            "5️⃣ Всё шифруется: Tor-шифрование, proxy, HTTPS."
        )
        await q.edit_message_text(
            about,
            parse_mode="HTML",
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("🔙 Назад", callback_data="back")],
                [InlineKeyboardButton("📤 Загрузить фото", callback_data="upload")]
            ])
        )
    elif data == "back":
        main_text = (
            "<b>👋 Конничива, бро и лёгкой руки!</b>\n\n"
            "Кидай мне фотку — я загружу на выбранный хостинг быстро и анонимно 🥷🏻."
        )
        await q.edit_message_text(
            main_text,
            parse_mode="HTML",
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("ℹ️ Почему мы?🥷🏻", callback_data="about")],
                [InlineKeyboardButton("📤 Загрузить фото 🥷🏻", callback_data="upload")]
            ])
        )

def main():
    ensure_workbook()
    app = Application.builder().token(TOKEN).build()

    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("help", help_command))
    app.add_handler(CallbackQueryHandler(button_handler))
    app.add_handler(MessageHandler(filters.PHOTO, on_photo))

    print("Bot is running…")
    app.run_polling()

if __name__ == "__main__":

    main()

